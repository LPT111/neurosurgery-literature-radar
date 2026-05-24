from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parents[1]
DEFAULT_OUTPUT = BASE_DIR / "config" / "journal_metrics.csv"


def normalize_journal_name(journal: str) -> str:
    return "".join(ch for ch in (journal or "").strip().lower() if ch.isalnum())


def clean(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def load_jcr(ws) -> dict[str, dict[str, str]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [clean(v) for v in rows[0]]
    index = {name: idx for idx, name in enumerate(header)}
    out: dict[str, dict[str, str]] = {}
    for row in rows[1:]:
        journal = clean(row[index.get("期刊名", 0)])
        if not journal:
            continue
        key = normalize_journal_name(journal)
        jif = clean(row[index.get("2024JIF", -1)]) if "2024JIF" in index else ""
        quartile = clean(row[index.get("Quartile", -1)]) if "Quartile" in index else ""
        rank = clean(row[index.get("JIF rank", -1)]) if "JIF rank" in index else ""
        category = clean(row[index.get("Category", -1)]) if "Category" in index else ""
        out[key] = {
            "journal": journal,
            "impact_factor": jif or "待核实",
            "jcr_quartile": quartile or "待核实",
            "jif_rank": rank,
            "category": category,
        }
    return out


def load_cas(ws) -> dict[str, dict[str, str]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}
    header = [clean(v) for v in rows[0]]
    index = {name: idx for idx, name in enumerate(header)}
    out: dict[str, dict[str, str]] = {}
    for row in rows[1:]:
        journal = clean(row[index.get("期刊名称", 0)])
        if not journal:
            continue
        key = normalize_journal_name(journal)
        cas = clean(row[index.get("2025分区", -1)]) if "2025分区" in index else ""
        top = clean(row[index.get("Top", -1)]) if "Top" in index else ""
        oa = clean(row[index.get("Open Access", -1)]) if "Open Access" in index else ""
        out[key] = {
            "journal": journal,
            "cas_quartile": f"CAS {cas}区" if cas else "待核实",
            "cas_top": top,
            "open_access": oa,
        }
    return out


def main() -> None:
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python scripts/import_journal_metrics.py /path/to/CAS_JCR.xlsx")
    input_path = Path(sys.argv[1]).expanduser()
    if not input_path.exists():
        raise SystemExit(f"File not found: {input_path}")

    wb = load_workbook(input_path, read_only=True, data_only=True)
    jcr_ws = wb["2025JCRIF-分区"]
    cas_ws = wb["2025中科学院分区表"]
    jcr = load_jcr(jcr_ws)
    cas = load_cas(cas_ws)

    keys = sorted(set(jcr) | set(cas), key=lambda k: (jcr.get(k) or cas.get(k) or {}).get("journal", k).lower())
    DEFAULT_OUTPUT.parent.mkdir(exist_ok=True)
    with DEFAULT_OUTPUT.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "journal",
                "impact_factor",
                "quartile",
                "jcr_quartile",
                "cas_quartile",
                "jif_rank",
                "category",
                "cas_top",
                "open_access",
                "source",
            ],
        )
        writer.writeheader()
        for key in keys:
            j = jcr.get(key, {})
            c = cas.get(key, {})
            journal = j.get("journal") or c.get("journal") or key
            impact_factor = j.get("impact_factor", "待核实")
            jcr_quartile = j.get("jcr_quartile", "待核实")
            cas_quartile = c.get("cas_quartile", "待核实")
            parts = []
            if jcr_quartile and jcr_quartile != "待核实":
                parts.append(f"JCR {jcr_quartile}" if not jcr_quartile.upper().startswith("JCR") else jcr_quartile)
            if cas_quartile and cas_quartile != "待核实":
                parts.append(cas_quartile)
            quartile = " / ".join(parts) if parts else "待核实"
            writer.writerow(
                {
                    "journal": journal,
                    "impact_factor": impact_factor,
                    "quartile": quartile,
                    "jcr_quartile": jcr_quartile,
                    "cas_quartile": cas_quartile,
                    "jif_rank": j.get("jif_rank", ""),
                    "category": j.get("category", ""),
                    "cas_top": c.get("cas_top", ""),
                    "open_access": c.get("open_access", ""),
                    "source": "2025 CAS/JCR user table",
                }
            )
    print(f"Wrote {len(keys)} journal metrics to {DEFAULT_OUTPUT}")


if __name__ == "__main__":
    main()
